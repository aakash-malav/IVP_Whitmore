"""
Whitmore Fund II Structured Credit Sleeve Dashboard
FinValley 10.0 Case Competition
Data loaded from Whitmore_Fund_II_Analysis.xlsx
Theme: professional finance (Indus Valley style navy / teal)
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path

st.set_page_config(
    page_title="Whitmore Fund II | Structured Credit Sleeve",
    page_icon="W",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Indus Valley style colours (brand accents: identical in both themes) ──
NAVY = "#0B1F3A"
TEAL = "#0D7377"
TEAL_LIGHT = "#14919B"
GOLD = "#C4A35A"
RED = "#B33A3A"
GREEN = "#2E7D4F"
AMBER = "#C47A00"
MUTED = "#5A6A7A"

# ═══════════════════════════════════════════════════════════
# THEME: one detection, one palette, used by CSS *and* Plotly
# ═══════════════════════════════════════════════════════════

# The override widget lives in the sidebar (created further down), but the CSS
# has to be injected before anything renders. Reading the widget's key straight
# out of session_state lets us know the user's choice on this run.
if "theme_mode" not in st.session_state:
    st.session_state["theme_mode"] = "Auto (follow browser)"


def detect_theme():
    """Return (mode, confident).

    mode is "dark" or "light". confident is False when we had to guess, in
    which case the CSS also emits a prefers-color-scheme fallback so the page
    still matches the browser even if Python could not tell.
    """
    choice = st.session_state.get("theme_mode", "Auto (follow browser)")
    if choice.startswith("Dark"):
        return "dark", True
    if choice.startswith("Light"):
        return "light", True

    # 1) st.context.theme.type — Streamlit >= 1.44. This resolves "Use system
    #    setting" against the real browser preference, so it is the one signal
    #    that actually tracks the OS/browser toggle.
    try:
        th = getattr(st.context, "theme", None)
        if th is not None:
            ttype = getattr(th, "type", None)
            if ttype is None and isinstance(th, dict):
                ttype = th.get("type")
            if ttype in ("dark", "light"):
                return ttype, True
    except Exception:
        pass

    # 2) Explicit base in config.toml
    try:
        base = st.get_option("theme.base")
        if base in ("dark", "light"):
            return base, True
    except Exception:
        pass

    # 3) Background luminance from config, if one was set
    try:
        bg = (st.get_option("theme.backgroundColor") or "").strip()
        if bg.startswith("#") and len(bg) >= 7:
            r, g, b = int(bg[1:3], 16), int(bg[3:5], 16), int(bg[5:7], 16)
            if (0.299 * r + 0.587 * g + 0.114 * b) < 128:
                return "dark", True
            return "light", True
    except Exception:
        pass

    return "light", False


THEME, THEME_CONFIDENT = detect_theme()
IS_DARK = THEME == "dark"

# Single palette. Every colour below is read by both the CSS and the charts,
# so page chrome and Plotly text can never disagree about which theme is live.
PALETTE = {
    "light": {
        "bg": "#F5F7FA",
        "card": "#FFFFFF",
        "text": "#0B1F3A",
        "muted": "#5A6A7A",
        "heading": "#0B1F3A",
        "border": "#D0D7DE",
        "grid": "rgba(11,31,58,0.12)",
        "axis": "#3C4C5E",
        "pie_label": "#FFFFFF",
        "ok_bg": "#E8F5E9", "ok_border": "#A5D6A7", "ok_text": "#1B5E20",
        "warn_bg": "#FFF8E1", "warn_border": "#FFE082", "warn_text": "#8A4B00",
        "info_bg": "#E3F2FD", "info_border": "#90CAF9", "info_text": "#0D47A1",
    },
    "dark": {
        "bg": "#0E1117",
        "card": "#161B22",
        "text": "#E8EEF4",
        "muted": "#9AA7B4",
        "heading": "#E8EEF4",
        "border": "#2A3038",
        "grid": "rgba(232,238,244,0.16)",
        "axis": "#C3CDD8",
        # Slice fills are the same saturated brand colours in both themes, so
        # the label sitting on top of them must not flip with the page theme.
        "pie_label": "#FFFFFF",
        "ok_bg": "#12301F", "ok_border": "#2E7D4F", "ok_text": "#A5D6A7",
        "warn_bg": "#332612", "warn_border": "#C47A00", "warn_text": "#FFD98A",
        "info_bg": "#12243A", "info_border": "#14919B", "info_text": "#9CD2F5",
    },
}
P = PALETTE[THEME]


def _theme_rules(p):
    """CSS rules for one palette. Reused for the media-query fallback."""
    return f"""
    .stApp {{ background-color: {p['bg']} !important; }}
    .stApp, .stApp p, .stApp li, .stApp label,
    [data-testid="stMarkdownContainer"] {{ color: {p['text']}; }}
    h1, h2, h3, h4, h5, h6 {{ color: {p['heading']} !important; }}
    [data-testid="stCaptionContainer"], .stCaption, small {{ color: {p['muted']} !important; }}
    [data-testid="stMetricValue"] {{ color: {p['text']} !important; font-weight: 700; }}
    [data-testid="stMetricLabel"], [data-testid="stMetricLabel"] * {{ color: {p['muted']} !important; }}
    [data-testid="stMetricDelta"] svg {{ fill: currentColor; }}
    hr {{ border-color: {p['border']} !important; }}
    [data-testid="stHeader"] {{ background: transparent !important; }}
    .finding-success {{ background: {p['ok_bg']}; border-color: {p['ok_border']}; color: {p['ok_text']}; }}
    .finding-warning {{ background: {p['warn_bg']}; border-color: {p['warn_border']}; color: {p['warn_text']}; }}
    .finding-info    {{ background: {p['info_bg']}; border-color: {p['info_border']}; color: {p['info_text']}; }}
    .finding-card .finding-title, .finding-card .finding-body,
    .finding-card strong {{ color: inherit !important; }}

    /* Plotly text, styled at the CSS layer as well as via the figure layout.
       st.context.theme can be one rerun stale right after a theme flip, so
       this keeps titles, ticks and legends readable in the meantime.
       .slicetext / .bartext are deliberately excluded: those sit on top of a
       coloured mark and get a per-slice colour computed in style_fig(). */
    .js-plotly-plot .gtitle,
    .js-plotly-plot .xtitle,
    .js-plotly-plot .ytitle,
    .js-plotly-plot .legendtext,
    .js-plotly-plot .legendtitletext,
    .js-plotly-plot .annotation-text {{ fill: {p['text']} !important; }}
    .js-plotly-plot .xtick text,
    .js-plotly-plot .ytick text,
    .js-plotly-plot .cbaxis .xtick text,
    .js-plotly-plot .cbaxis .ytick text {{ fill: {p['axis']} !important; }}
"""


# In Auto mode the browser's own prefers-color-scheme is the fastest and most
# accurate source, so the base rules are light and a media query flips them.
# When the user pins Light or Dark, that choice wins and no query is emitted.
AUTO_THEME = str(st.session_state.get("theme_mode", "")).startswith("Auto")
CSS_BASE = PALETTE["light"] if (AUTO_THEME or not THEME_CONFIDENT) else P

CUSTOM_CSS = f"""
<style>
    /* ── Palette resolved in Python: charts and chrome share these values ── */
{_theme_rules(CSS_BASE)}

    /* Sidebar is always navy in both themes — brand constant */
    [data-testid="stSidebar"] {{ background-color: {NAVY} !important; }}
    [data-testid="stSidebar"] * {{ color: #E8EEF4 !important; }}
    [data-testid="stSidebar"] [data-testid="stMetricValue"] {{ color: #FFFFFF !important; }}
    [data-testid="stSidebar"] [data-testid="stCaptionContainer"] {{ color: #B9C6D4 !important; }}

    .block-container {{ padding-top: 2.25rem !important; padding-bottom: 2rem; }}

    /* Decision banner – gradient works on either background */
    .decision-banner {{
        background: linear-gradient(90deg, #0B1F3A 0%, #0D7377 100%) !important;
        color: #FFFFFF !important;
        padding: 1rem 1.35rem !important;
        border-radius: 10px !important;
        font-size: 1.08rem !important;
        font-weight: 650 !important;
        margin: 0.75rem 0 1.25rem 0 !important;
        line-height: 1.5 !important;
        display: block !important;
        width: 100% !important;
        box-sizing: border-box !important;
        border: 1px solid rgba(255,255,255,0.2) !important;
        box-shadow: 0 2px 10px rgba(0,0,0,0.25) !important;
        position: relative !important;
        z-index: 5 !important;
    }}
    .decision-banner * {{ color: #FFFFFF !important; }}

    /* Finding cards – fixed layout so text never collapses */
    .finding-card {{
        border-radius: 10px;
        padding: 1rem 1.1rem;
        height: 100%;
        min-height: 150px;
        display: flex;
        flex-direction: column;
        gap: 0.5rem;
        border: 1px solid transparent;
        box-sizing: border-box;
    }}
    .finding-card .finding-title {{
        font-weight: 700;
        font-size: 0.95rem;
        line-height: 1.3;
        margin: 0 0 0.35rem 0;
    }}
    .finding-card .finding-body {{
        font-size: 0.88rem;
        line-height: 1.5;
        margin: 0;
        opacity: 0.95;
    }}
</style>
"""

# Emitted whenever the browser is the authority on the theme: Auto mode, or an
# older Streamlit where Python could not read the theme at all.
if AUTO_THEME or not THEME_CONFIDENT:
    CUSTOM_CSS += f"""
<style>
@media (prefers-color-scheme: dark) {{
{_theme_rules(PALETTE['dark'])}
}}
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


def load_metrics_from_excel(path: str):
    """Read Performance_Metrics with openpyxl data_only so values are numbers."""
    from openpyxl import load_workbook
    defaults = {
        "funded": 456475000,
        "cash_int": 9967750,
        "pik": 781250,
        "loan_income": 10749000,
        "cost_debt": 13593750,
        "net_income": -2844750,
        "timing_gap": 19945085,
        "orion_wd": 13125000,
        "orion_cds": 14437500,
        "orion_net": 1312500,
        "irs_mtm": -3842000,
        "irs_basis": 367825,
        "yield_pre": 0.0942,
        "yield_post": -0.0249,
        "gap_pct": 0.0437,
        "hedge": "Imperfect",
        "longwood": 1800000,
        "cash_conv": 0.9273,
        "pik_share": 0.0727,
        "irs_bps_book": 8.06,
        "irs_days": 3.08,
        "jpy_residual": 2150000,
        "jpy_bps": 47.1,
        "combined_residual": 2517825,
        "combined_bps": 55.16,
        "be_yield": 0.1191,
        "yield_gap_be": -0.0249,
        "lev_drag": 0.1191,
        "shortfall": -2844750,
    }
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return defaults
    if "Performance_Metrics" not in wb.sheetnames:
        return defaults
    ws = wb["Performance_Metrics"]
    mapping = {
        "funded": "total funded balance",
        "cash_int": "cash interest received",
        "pik": "accrued pik",
        "loan_income": "total loan income",
        "cost_debt": "cost of debt",
        "net_income": "net income before",
        "timing_gap": "recognition timing gap (absolute)",
        "orion_wd": "orion mezzanine write-down",
        "orion_cds": "orion cds settlement",
        "orion_net": "net orion impact",
        "irs_mtm": "irs mark-to-market",
        "irs_basis": "irs basis mismatch",
        "yield_pre": "net yield before leverage",
        "yield_post": "net yield after debt cost",
        "gap_pct": "recognition timing gap as %",
        "longwood": "longwood lease cost",
        "cash_conv": "cash conversion ratio",
        "pik_share": "pik share of loan income",
        "irs_bps_book": "irs basis drag: bps",
        "irs_days": "irs basis drag: days",
        "jpy_residual": "jpy residual fx exposure",
        "jpy_bps": "jpy residual: bps",
        "combined_residual": "combined hedge residual cost",
        "combined_bps": "combined residual: bps",
        "be_yield": "break-even asset yield",
        "yield_gap_be": "yield gap vs break-even",
        "lev_drag": "leverage drag",
        "shortfall": "quarterly income shortfall",
    }
    out = defaults.copy()
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 2).value  # column B = Metric
        if label is None:
            continue
        label_l = str(label).strip().lower()
        val = ws.cell(r, 3).value   # column C = Value
        for key, needle in mapping.items():
            if needle in label_l:
                if key == "hedge":
                    continue
                try:
                    num = float(val)
                    if key == "cost_debt":
                        out[key] = abs(num)
                    else:
                        out[key] = num
                except (TypeError, ValueError):
                    if key == "hedge" or "imperfect" in str(val).lower():
                        out["hedge"] = str(val).strip()
                break
        if "hedge effectiveness" in label_l and val is not None:
            out["hedge"] = str(val).strip()
    return out


@st.cache_data
def load_workbook(path: str):
    xls = pd.ExcelFile(path)
    sheets = {}
    for name in xls.sheet_names:
        try:
            sheets[name] = pd.read_excel(xls, sheet_name=name, header=None)
        except Exception:
            pass
    return sheets


def extract_performance(sheets):
    # Kept for compatibility; real load uses load_metrics_from_excel
    return {
        "funded": 456475000,
        "cash_int": 9967750,
        "pik": 781250,
        "loan_income": 10749000,
        "cost_debt": 13593750,
        "net_income": -2844750,
        "timing_gap": 19945085,
        "orion_wd": 13125000,
        "orion_cds": 14437500,
        "orion_net": 1312500,
        "irs_mtm": -3842000,
        "irs_basis": 367825,
        "yield_pre": 0.0942,
        "yield_post": -0.0249,
        "gap_pct": 0.0437,
        "hedge": "Imperfect",
        "longwood": 1800000,
    }


def extract_loan_book(sheets):
    if "Loan_Book" not in sheets:
        return pd.DataFrame()
    raw = sheets["Loan_Book"]
    # Find header row containing Loan_ID
    header_idx = None
    for i in range(min(15, len(raw))):
        row = [str(x).strip() for x in raw.iloc[i].tolist()]
        if any("Loan_ID" in x for x in row):
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()
    headers = [str(x).strip() if pd.notna(x) else f"col{j}" for j, x in enumerate(raw.iloc[header_idx])]
    data = raw.iloc[header_idx + 1:].copy()
    data.columns = headers
    data = data.dropna(how="all")
    # Keep useful columns
    keep = {}
    for h in headers:
        hl = h.lower()
        if "loan_id" in hl:
            keep["Loan_ID"] = h
        elif "obligor" in hl:
            keep["Obligor"] = h
        elif "industry" in hl:
            keep["Industry"] = h
        elif "tranche" in hl:
            keep["Tranche"] = h
        elif "current_balance" in hl or "current balance" in hl:
            keep["Balance"] = h
        elif "status" in hl:
            keep["Status"] = h
        elif "cash_interest" in hl or "cash interest" in hl:
            keep["Cash_Int"] = h
        elif "accrued_pik" in hl or "accrued pik" in hl:
            keep["PIK"] = h
    if not keep:
        return pd.DataFrame()
    out = data[[keep[k] for k in keep]].copy()
    out.columns = list(keep.keys())
    for col in ["Balance", "Cash_Int", "PIK"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)
    out = out[out["Balance"] > 0] if "Balance" in out.columns else out
    return out.reset_index(drop=True)


def extract_decisions(sheets):
    if "Decision_Quality" not in sheets:
        return pd.DataFrame()
    raw = sheets["Decision_Quality"]
    header_idx = None
    for i in range(min(20, len(raw))):
        row = [str(x).strip() for x in raw.iloc[i].tolist()]
        if any("Decision_ID" in x or "Decision_Area" in x for x in row):
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame([
            {"ID": "D-01", "Area": "Orion CDS Direction", "Conf": 9, "Status": "Closed"},
            {"ID": "D-04", "Area": "IRS Effectiveness", "Conf": 9, "Status": "Open"},
            {"ID": "D-07", "Area": "JPY Forward Notional", "Conf": 6, "Status": "Open"},
        ])
    headers = [str(x).strip() if pd.notna(x) else f"c{j}" for j, x in enumerate(raw.iloc[header_idx])]
    data = raw.iloc[header_idx + 1:].copy()
    data.columns = headers
    data = data.dropna(how="all")
    cols = {}
    for h in headers:
        hl = h.lower()
        if "decision_id" in hl or h == "Decision_ID":
            cols["ID"] = h
        elif "decision_area" in hl or "area" in hl:
            cols["Area"] = h
        elif "confidence_score" in hl or "score" in hl:
            cols["Conf"] = h
        elif "status" in hl:
            cols["Status"] = h
    if not cols:
        return pd.DataFrame()
    out = data[[cols[k] for k in cols]].copy()
    out.columns = list(cols.keys())
    if "Conf" in out.columns:
        out["Conf"] = pd.to_numeric(out["Conf"], errors="coerce")
    return out.dropna(subset=["ID"]).reset_index(drop=True)


# Load data
_candidates = [
    Path("/home/workdir/attachments/Whitmore_Fund_II_Analysis.xlsx"),
    Path(__file__).parent / "Whitmore_Fund_II_Analysis.xlsx",
    Path("/home/workdir/artifacts/Whitmore_Fund_II_Analysis_EXTENDED.xlsx"),
    Path("Whitmore_Fund_II_Analysis.xlsx"),
    Path("D:/Case_Comp_Guide/IVP/Whitmore_Fund_II_Analysis.xlsx"),
]
XLSX_PATH = next((p for p in _candidates if p.exists()), _candidates[0])

sheets = load_workbook(str(XLSX_PATH)) if XLSX_PATH.exists() else {}
M = load_metrics_from_excel(str(XLSX_PATH)) if XLSX_PATH.exists() else extract_performance({})
loans = extract_loan_book(sheets)
decisions = extract_decisions(sheets)

# ── Sidebar ──
st.sidebar.markdown(f"<h2 style='color:#E8EEF4;margin-bottom:0'>WHITMORE FUND II</h2>", unsafe_allow_html=True)
st.sidebar.caption("Structured Credit Sleeve | Q4 2025")
st.sidebar.markdown("---")

page = st.sidebar.radio(
    "Who are you?",
    [
        "Start here",
        "IC / Senior Management",
        "Portfolio Manager",
        "Fund Accounting",
        "Risk",
        "Operations",
        "LP / Investor",
        "Scenario Lab",
        "Adjustment Register",
    ],
)

st.sidebar.markdown("---")
st.sidebar.subheader("Scenario levers")
st.sidebar.caption("These levers recompute yields on this page. Underlying Excel remains the source of truth.")

sofr_bps = st.sidebar.slider("SOFR shock (bps)", -100, 200, 0, 25)
pik_extra = st.sidebar.checkbox("Two extra loans switch to PIK", value=False)
util_up = st.sidebar.checkbox("Greystone utilisation rises to 80%", value=False)

# Scenario engine (simple, transparent)
# Floating share ~65%; facility full sensitivity
float_share = 0.65
loan_income_s = M["loan_income"] * (1 + float_share * (sofr_bps / 10000))
cost_debt_s = M["cost_debt"] * (1 + 1.0 * (sofr_bps / 10000))
if pik_extra:
    loan_income_s -= 1_200_000
if util_up:
    cost_debt_s += 606_250
net_s = loan_income_s - cost_debt_s
yield_post_s = (net_s / M["funded"]) * 4
yield_pre_s = (loan_income_s / M["funded"]) * 4

st.sidebar.markdown("---")
st.sidebar.metric("Scenario yield pre-lev", f"{yield_pre_s:.2%}")
st.sidebar.metric("Scenario yield post-lev", f"{yield_post_s:.2%}")
if sofr_bps != 0 or pik_extra or util_up:
    st.sidebar.warning("Scenario active. Compare to base case on Scenario Lab page.")
else:
    st.sidebar.success("Base case (no shocks)")

st.sidebar.markdown("---")
# Bound to session_state["theme_mode"], which detect_theme() reads at the top of
# the script. Changing it reruns the app, so page CSS and charts flip together.
st.sidebar.radio(
    "Appearance",
    ["Auto (follow browser)", "Light", "Dark"],
    key="theme_mode",
    help="Auto follows your browser / OS light-dark setting. Pick Light or Dark to pin it.",
)
st.sidebar.caption(f"Active theme: {THEME.capitalize()}")
st.sidebar.caption("Data source: Whitmore_Fund_II_Analysis.xlsx")
st.sidebar.caption("Confidential | FinValley 10.0")


def kpi_strip(metrics_override=None):
    m = metrics_override or M
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Funded Balance", f"${m['funded']/1e6:.1f}m")
    c2.metric("Yield Pre-Lev", f"{m['yield_pre']:.2%}")
    c3.metric("Yield Post-Lev", f"{m['yield_post']:.2%}")
    c4.metric("Timing Gap", f"${m['timing_gap']/1e6:.1f}m")
    c5.metric("Orion Net", f"${m['orion_net']/1e6:.2f}m")
    c6.metric("Hedge", str(m.get("hedge", "Imperfect"))[:12])


def is_dark_theme() -> bool:
    """Kept for compatibility. THEME is resolved once, at the top of the script."""
    return IS_DARK


def _readable_on(fill) -> str:
    """Near-black or near-white, whichever reads better on `fill`."""
    try:
        h = str(fill).strip()
        if h.startswith("rgb"):
            r, g, b = [int(float(x)) for x in h[h.index("(") + 1:h.index(")")].split(",")[:3]]
        else:
            h = h.lstrip("#")
            if len(h) == 3:
                h = "".join(c * 2 for c in h)
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except Exception:
        return "#FFFFFF"

    def _lin(c):
        c /= 255.0
        return c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4

    lum = 0.2126 * _lin(r) + 0.7152 * _lin(g) + 0.0722 * _lin(b)
    on_white = 1.05 / (lum + 0.05)              # contrast vs #FFFFFF
    on_black = (lum + 0.05) / (0.00572 + 0.05)  # contrast vs #111111
    return "#FFFFFF" if on_white >= on_black else "#111111"


def chart_layout(**kwargs):
    """Plotly layout defaults driven by the same palette the CSS uses."""
    text = P["text"]
    axis = P["axis"]
    grid = P["grid"]
    base = dict(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color=text, size=13),
        title=dict(font=dict(color=text, size=15)),
        legend=dict(font=dict(color=text), bgcolor="rgba(0,0,0,0)"),
        hoverlabel=dict(
            bgcolor=P["card"],
            font=dict(color=text),
            bordercolor=P["border"],
        ),
        xaxis=dict(
            gridcolor=grid,
            zerolinecolor=grid,
            linecolor=grid,
            color=axis,
            title_font=dict(color=text),
            tickfont=dict(color=axis),
        ),
        yaxis=dict(
            gridcolor=grid,
            zerolinecolor=grid,
            linecolor=grid,
            color=axis,
            title_font=dict(color=text),
            tickfont=dict(color=axis),
        ),
        margin=dict(l=40, r=20, t=50, b=40),
    )
    if "title" in kwargs and isinstance(kwargs["title"], str):
        kwargs["title"] = dict(text=kwargs["title"], font=dict(color=text, size=15))
    base.update(kwargs)
    return base


def style_fig(fig, **layout_kwargs):
    """Apply the palette and force contrast on every text element Plotly draws.

    Plotly Express builds its own title/legend/colourbar fonts, so each of these
    has to be overridden explicitly or it falls back to Plotly's dark-on-light
    default and disappears against a dark background.
    """
    text = P["text"]
    axis = P["axis"]
    fig.update_layout(**chart_layout(**layout_kwargs))
    fig.update_layout(
        title_font_color=text,
        font_color=text,
        legend_font_color=text,
        legend_title_font_color=text,
    )
    fig.update_xaxes(tickfont_color=axis, title_font_color=text, color=axis)
    fig.update_yaxes(tickfont_color=axis, title_font_color=text, color=axis)

    # Labels inside a slice must contrast with the *fill*, which is the same
    # brand colour in both themes. So pick per-slice from the fill's luminance
    # rather than from the page theme (white on gold would fail either way).
    try:
        for tr in fig.data:
            if tr.type != "pie":
                continue
            fills = getattr(getattr(tr, "marker", None), "colors", None)
            if fills:
                colors = [_readable_on(c) for c in fills]
            else:
                colors = P["pie_label"]
            tr.update(
                insidetextfont_color=colors,
                outsidetextfont_color=text,
            )
    except Exception:
        fig.update_traces(insidetextfont_color=P["pie_label"], selector=dict(type="pie"))

    # Bar/waterfall labels drawn outside the mark follow page text.
    for ttype in ("bar", "waterfall"):
        try:
            fig.update_traces(outsidetextfont_color=text, selector=dict(type=ttype))
        except Exception:
            pass

    # Continuous colour scales carry their own tick + title fonts.
    try:
        fig.update_coloraxes(
            colorbar_tickfont_color=axis,
            colorbar_title_font_color=text,
            colorbar_outlinecolor=P["border"],
        )
    except Exception:
        pass
    return fig


# ═══════════════════════════════════════════════════════════
# PAGES
# ═══════════════════════════════════════════════════════════

if page == "Start here":
    st.title("Whitmore Fund II")
    st.subheader("Structured Credit Sleeve | Q4 2025")
    st.caption("17 loan positions | Data reconciled from Alpha, Beta, notices and confirmations")
    st.markdown("---")

    st.markdown("### What this is, in thirty seconds")
    st.info(
        "This sleeve lends to entertainment companies and runs a book of hedges. "
        "Source files from servicers, brokers and counterparties disagreed with each other. "
        "This dashboard shows numbers after those conflicts are resolved using a clear source hierarchy, "
        "and every material judgement is logged with confidence and impact.\n\n"
        "**Takeaway:** Asset level yield is healthy at about 9.4%. "
        "After debt cost the sleeve is currently negative. "
        "Both statements are true. Reporting only one of them is incomplete."
    )

    st.markdown("### Four findings that matter most")
    a, b, c, d = st.columns(4)
    with a:
        st.markdown(
            """
<div class="finding-card finding-success">
<p class="finding-title">Orion CDS is protection bought</p>
<p class="finding-body">Settlement $14.4m received.<br>Net impact after write-down: <strong>+$1.31m</strong>.</p>
</div>
            """,
            unsafe_allow_html=True,
        )
    with b:
        st.markdown(
            """
<div class="finding-card finding-warning">
<p class="finding-title">IRS hedge is imperfect</p>
<p class="finding-body">Permanent 14.7 bps basis.<br>Ongoing cost about <strong>$368k</strong> per year.</p>
</div>
            """,
            unsafe_allow_html=True,
        )
    with c:
        st.markdown(
            """
<div class="finding-card finding-warning">
<p class="finding-title">JPY forward under-hedged</p>
<p class="finding-body">Loan notional reset in Nov 2025.<br>Forward not updated. Residual about <strong>$2.15m</strong>.</p>
</div>
            """,
            unsafe_allow_html=True,
        )
    with d:
        st.markdown(
            """
<div class="finding-card finding-info">
<p class="finding-title">About 7% of Q4 income is PIK</p>
<p class="finding-body"><strong>$781k</strong> recognised, not cash this quarter.</p>
</div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("### Pick your seat — each view answers one decision")
    n1, n2, n3 = st.columns(3)
    n1.markdown("**IC / Senior Mgmt**  \nKeep, fix, or scale the sleeve?")
    n2.markdown("**Portfolio Manager**  \nDeploy, hedge, restructure?")
    n3.markdown("**Fund Accounting**  \nCan we defend these numbers?")
    n4, n5, n6 = st.columns(3)
    n4.markdown("**Risk**  \nWhat residual risk remains?")
    n5.markdown("**Operations**  \nWhat do we fix this week?")
    n6.markdown("**LP / Investor**  \nStay, redeem, or commit more?")

elif page == "IC / Senior Management":
    st.markdown(
        '<div class="decision-banner">Decision: Keep · Fix · or Scale this sleeve?</div>',
        unsafe_allow_html=True,
    )
    st.title("Investment Committee View")
    st.caption("One-page health for IC and senior management")
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Asset Yield", f"{M['yield_pre']:.2%}")
    c2.metric("Levered Yield", f"{M['yield_post']:.2%}")
    c3.metric("Break-even Yield", f"{M.get('be_yield', 0.1191):.2%}")
    c4.metric("Gap vs BE", f"{M.get('yield_gap_be', M['yield_post']):+.2%}")
    c5.metric("Orion Net", f"+${M['orion_net']/1e6:.2f}m")
    c6.metric("Residual Risk", f"{M.get('combined_bps', 55):.0f} bps")
    st.markdown("---")

    col1, col2, col3 = st.columns(3)
    with col1:
        fig = go.Figure(data=[
            go.Bar(name="Loan Income", x=["Q4 2025"], y=[M["loan_income"]], marker_color=TEAL),
            go.Bar(name="Cost of Debt", x=["Q4 2025"], y=[M["cost_debt"]], marker_color=RED),
        ])
        style_fig(fig, title="Income vs Cost of Debt", barmode="group", height=340)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        fig2 = go.Figure(data=[
            go.Bar(name="Write-down", x=["Orion"], y=[-M["orion_wd"]], marker_color=RED),
            go.Bar(name="CDS Gain", x=["Orion"], y=[M["orion_cds"]], marker_color=GREEN),
            go.Bar(name="Net", x=["Orion"], y=[M["orion_net"]], marker_color=TEAL),
        ])
        style_fig(fig2, title="Orion Restructuring Impact", barmode="group", height=340)
        st.plotly_chart(fig2, use_container_width=True)

    with col3:
        fig3 = go.Figure(data=[
            go.Bar(name="Yield",
                   x=["Pre-lev", "Break-even", "Post-lev"],
                   y=[M["yield_pre"] * 100, M.get("be_yield", 0.1191) * 100, M["yield_post"] * 100],
                   marker_color=[GREEN, AMBER, RED]),
        ])
        style_fig(fig3, title="Yield vs Break-even (ann. %)", height=340, yaxis_title="%")
        st.plotly_chart(fig3, use_container_width=True)

    gap_df = pd.DataFrame({
        "Component": ["Orion Write-down", "IRS MTM Gap", "FX Realised Gap", "PIK + Missed"],
        "USD": [M["orion_wd"], 3367000, 2671460, 781625],
    })
    col4, col5 = st.columns(2)
    with col4:
        fig4 = px.pie(gap_df, values="USD", names="Component", title="Timing Gap Composition",
                      color_discrete_sequence=[RED, AMBER, TEAL, GOLD])
        style_fig(fig4, height=360)
        st.plotly_chart(fig4, use_container_width=True)
    with col5:
        st.markdown("### Key insights")
        st.markdown(
            f"""
- Asset yield is healthy at **{M['yield_pre']:.2%}**.
- After debt cost the sleeve is at **{M['yield_post']:.2%}**.
- Timing gap is **${M['timing_gap']/1e6:.1f}m** ({M['gap_pct']:.2%} of book).
- Orion net impact is **+${M['orion_net']/1e6:.2f}m** (CDS recovery exceeded write-down).
- IRS basis leakage is about **${M['irs_basis']:,.0f}** per year.
- Two open items remain: IRS restrike and JPY forward notional update.
"""
        )

elif page == "Portfolio Manager":
    st.markdown(
        '<div class="decision-banner">Decision: Deploy capital · Hedge · or Restructure positions?</div>',
        unsafe_allow_html=True,
    )
    st.title("Portfolio Manager View")
    st.caption("Yield quality, income mix, loan book, open hedge actions")
    p1, p2, p3, p4, p5, p6 = st.columns(6)
    p1.metric("Yield Pre", f"{M['yield_pre']:.2%}")
    p2.metric("Yield Post", f"{M['yield_post']:.2%}")
    p3.metric("Cash Conversion", f"{M.get('cash_conv', 0.927):.1%}")
    p4.metric("PIK Share", f"{M.get('pik_share', 0.073):.1%}")
    p5.metric("Orion Net", f"+${M['orion_net']/1e6:.2f}m")
    p6.metric("Hedge", str(M.get("hedge", "Imperfect"))[:12])
    st.markdown("---")
    if False:
        kpi_strip()
    st.markdown("---")

    c1, c2 = st.columns(2)
    with c1:
        mix = pd.DataFrame({"Type": ["Cash Interest", "PIK Accrued"], "USD": [M["cash_int"], M["pik"]]})
        fig = px.pie(mix, values="USD", names="Type", title="Cash vs PIK Income (Q4)",
                     color_discrete_sequence=[TEAL, GOLD])
        style_fig(fig, height=350)
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        if not loans.empty and "Industry" in loans.columns:
            ind = loans.groupby("Industry", as_index=False)["Balance"].sum().sort_values("Balance")
            fig2 = px.bar(ind, x="Balance", y="Industry", orientation="h", title="Funded Balance by Industry",
                          color_discrete_sequence=[TEAL])
            style_fig(fig2, height=350)
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Loan book industry chart requires Loan_Book sheet.")

    if not loans.empty:
        st.subheader("Loan book (from Excel)")
        show = loans.copy()
        if "Balance" in show.columns:
            show = show.sort_values("Balance", ascending=False)
        st.dataframe(
            show.style.format({c: "${:,.0f}" for c in ["Balance", "Cash_Int", "PIK"] if c in show.columns}),
            use_container_width=True, hide_index=True,
        )
        if "Status" in loans.columns:
            status_counts = loans["Status"].value_counts().reset_index()
            status_counts.columns = ["Status", "Count"]
            fig3 = px.bar(status_counts, x="Status", y="Count", title="Loans by Status",
                          color="Status", color_discrete_sequence=[TEAL, AMBER, GOLD, RED])
            style_fig(fig3, height=300)
            st.plotly_chart(fig3, use_container_width=True)

elif page == "Fund Accounting":
    st.markdown(
        '<div class="decision-banner">Decision: Can we close the books and defend these numbers?</div>',
        unsafe_allow_html=True,
    )
    st.title("Fund Accounting View")
    st.caption("Cash vs recognised · adjustments · PIK · source hierarchy")
    a1, a2, a3, a4, a5, a6 = st.columns(6)
    a1.metric("Cash Interest", f"${M['cash_int']/1e6:.2f}m")
    a2.metric("Recognised Income", f"${M['loan_income']/1e6:.2f}m")
    a3.metric("Cash Conversion", f"{M.get('cash_conv', 0.927):.1%}")
    a4.metric("Timing Gap", f"${M['timing_gap']/1e6:.1f}m")
    a5.metric("Gap % Book", f"{M['gap_pct']:.2%}")
    a6.metric("PIK (Q4)", f"${M['pik']/1e3:.0f}k")
    st.markdown("---")
    st.subheader("Recognition bridge — cash to recognised")

    bridge = pd.DataFrame([
        {"Line": "Cash Interest (performing)", "Cash": M["cash_int"], "Recognised": M["cash_int"], "Gap": 0},
        {"Line": "PIK Interest (Emberlight)", "Cash": 0, "Recognised": M["pik"], "Gap": M["pik"]},
        {"Line": "Orion Mezz Write-down", "Cash": 0, "Recognised": -M["orion_wd"], "Gap": M["orion_wd"]},
        {"Line": "Orion CDS Settlement", "Cash": M["orion_cds"], "Recognised": M["orion_cds"], "Gap": 0},
        {"Line": "IRS MTM vs cash paid", "Cash": -475000, "Recognised": M["irs_mtm"], "Gap": 3367000},
        {"Line": "FX Forwards Realised", "Cash": -96000, "Recognised": -2767460, "Gap": 2671460},
        {"Line": "Debt Interest Paid", "Cash": -M["cost_debt"], "Recognised": -M["cost_debt"], "Gap": 0},
    ])
    st.dataframe(
        bridge.style.format({"Cash": "${:,.0f}", "Recognised": "${:,.0f}", "Gap": "${:,.0f}"}),
        use_container_width=True, hide_index=True,
    )

    gap_df = pd.DataFrame({
        "Component": ["Orion Write-down", "IRS MTM Gap", "FX Realised Gap", "PIK + Missed"],
        "USD": [M["orion_wd"], 3367000, 2671460, 781625],
    })
    c1, c2 = st.columns(2)
    with c1:
        fig = px.bar(gap_df, x="Component", y="USD", title="Timing Gap Drivers",
                     color="USD", color_continuous_scale=["#F5B7B1", RED])
        style_fig(fig, height=380)
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        fig2 = go.Figure(go.Waterfall(
            name="Gap", orientation="v",
            measure=["relative", "relative", "relative", "relative", "total"],
            x=["Orion WD", "IRS MTM", "FX Realised", "PIK", "Total Gap"],
            y=[M["orion_wd"], 3367000, 2671460, 781625, 0],
            connector={"line": {"color": P["axis"]}},
            increasing={"marker": {"color": RED}},
            totals={"marker": {"color": TEAL_LIGHT if IS_DARK else NAVY}},
        ))
        style_fig(fig2, title="Waterfall of Absolute Gaps", height=380)
        st.plotly_chart(fig2, use_container_width=True)

    st.warning(
        f"Total recognition timing gap: **${M['timing_gap']:,.0f}** "
        f"({M['gap_pct']:.2%} of funded balance). "
        "A single gap number hides settlement lag, PIK that may not convert, write-downs with no cash, and MTM changes."
    )

elif page == "Risk":
    st.markdown(
        '<div class="decision-banner">Decision: What residual risk remains — and is it within appetite?</div>',
        unsafe_allow_html=True,
    )
    st.title("Risk View")
    st.caption("Timing gap · concentration · hedge residual · open items")

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Timing Gap", f"${M['timing_gap']/1e6:.1f}m")
    c2.metric("Gap % Book", f"{M['gap_pct']:.2%}")
    c3.metric("IRS Basis", f"{M.get('irs_bps_book', 8.06):.1f} bps")
    c4.metric("JPY Residual", f"{M.get('jpy_bps', 47.1):.0f} bps")
    c5.metric("Combined Residual", f"{M.get('combined_bps', 55.2):.0f} bps")
    c6.metric("Top Name", "26.3% Lumivue")
    st.markdown("---")
    open_n = int((decisions["Status"].astype(str).str.contains("Open", case=False)).sum()) if not decisions.empty and "Status" in decisions.columns else 2
    high_n = int((decisions["Conf"] >= 8).sum()) if not decisions.empty and "Conf" in decisions.columns else 6
    total_d = len(decisions) if not decisions.empty else 8

    c1, c2 = st.columns(2)
    with c1:
        gap_df = pd.DataFrame({
            "Component": ["Orion Write-down", "IRS MTM", "FX Realised", "PIK + Missed"],
            "USD": [M["orion_wd"], 3367000, 2671460, 781625],
        })
        fig = px.bar(gap_df, x="USD", y="Component", orientation="h", title="Gap by Source",
                     color_discrete_sequence=[TEAL])
        style_fig(fig, height=360)
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        st.subheader("Decision Quality Log")
        if not decisions.empty:
            st.dataframe(decisions, use_container_width=True, hide_index=True)
        else:
            st.write("Load Decision_Quality sheet for full log.")

    st.markdown("### Counterparty dual-channel: Greystone")
    g1, g2, g3, g4 = st.columns(4)
    g1.metric("Bridge loan (asset)", "$12.0m")
    g2.metric("IRS notional", "$250m")
    g3.metric("Q4 facility interest", "$5.25m")
    g4.metric("Headline touchpoint", "$262m")
    st.caption("Greystone is both a loan obligor and the dominant financing/hedge counterparty — multi-channel concentration.")

    st.error(
        "**Open action items**\n\n"
        f"1. IRS basis risk: permanent 14.7 bps mismatch. Annual cost about ${M['irs_basis']:,.0f} "
        f"({M.get('irs_bps_book', 8.06):.1f} bps of book). Action: restrike or unwind.\n\n"
        f"2. JPY forward under-hedge: residual ~${M.get('jpy_residual', 2150000)/1e6:.2f}m "
        f"({M.get('jpy_bps', 47.1):.0f} bps of book). Action: re-size forward."
    )

elif page == "Adjustment Register":
    st.title("Adjustment Register")
    st.caption("Material judgements with source hierarchy rank")
    st.markdown(
        "**Rank 1** = Contract / ISDA confirmation &nbsp;|&nbsp; "
        "**Rank 2** = Formal notice &nbsp;|&nbsp; "
        "**Rank 3** = Servicer tape &nbsp;|&nbsp; "
        "**Rank 4** = PM / ops note"
    )
    adj = pd.DataFrame([
        {"R_ID": "R01", "Rank": 1, "Target": "Orion CDS", "Type": "Direction", "Amount": M["orion_cds"], "Status": "Closed", "Source": "Main Confirmation: Protection Buyer"},
        {"R_ID": "R02", "Rank": 2, "Target": "Orion Mezz", "Type": "Impairment", "Amount": -M["orion_wd"], "Status": "Closed", "Source": "RSA 15-Aug-2025: 47.5% of face"},
        {"R_ID": "R03", "Rank": 2, "Target": "Timberline", "Type": "Balance correction", "Amount": -4500000, "Status": "Closed", "Source": "Notice 3: $4.5m prepayment"},
        {"R_ID": "R04", "Rank": 2, "Target": "Emberlight PIK", "Type": "PIK accrual", "Amount": 1850000, "Status": "Closed", "Source": "Notice: PIK election confirmed"},
        {"R_ID": "R05", "Rank": 1, "Target": "IRS", "Type": "Effectiveness", "Amount": -M["irs_basis"], "Status": "Open", "Source": "ISDA Fallback + facility CSA"},
        {"R_ID": "R06", "Rank": 2, "Target": "JPY FX", "Type": "Under-hedge", "Amount": -2150000, "Status": "Open", "Source": "Notice 4: notional reset"},
        {"R_ID": "R07", "Rank": 1, "Target": "Longwood Lease", "Type": "Policy", "Amount": -M["longwood"], "Status": "Closed", "Source": "Portfolio view, LAIM strategy"},
        {"R_ID": "R08", "Rank": 3, "Target": "Duplicates", "Type": "De-duplication", "Amount": 0, "Status": "Closed", "Source": "Case rule + transfer dates"},
    ])
    st.dataframe(adj.style.format({"Amount": "${:,.0f}"}), use_container_width=True, hide_index=True)

    rank_counts = adj["Rank"].value_counts().sort_index().reset_index()
    rank_counts.columns = ["Rank", "Count"]
    fig = px.bar(rank_counts, x="Rank", y="Count", title="Adjustments by Source Rank",
                 color="Rank", color_continuous_scale=["#0D7377", "#14919B", "#C4A35A", "#B33A3A"])
    style_fig(fig, height=320)
    st.plotly_chart(fig, use_container_width=True)

    st.info(
        "Rank 1 governs absolutely. Rank 2 supersedes system state. "
        "Ranks 3 and 4 are used only when higher ranks are silent. "
        "This register makes every material judgement audit ready."
    )

elif page == "LP / Investor":
    st.markdown(
        '<div class="decision-banner">Decision: Stay · Redeem · or Commit more capital?</div>',
        unsafe_allow_html=True,
    )
    st.title("LP / Investor View")
    st.caption("Returns, capital flows, risk flags")

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Yield after Debt", f"{M['yield_post']:.2%}")
    c2.metric("Yield Pre-Lev", f"{M['yield_pre']:.2%}")
    c3.metric("Capital Called Q4", "$25.0m")
    c4.metric("Distributions Q4", "$15.0m")
    c5.metric("Net Capital In", "+$6.8m")
    c6.metric("Cash Conversion", f"{M.get('cash_conv', 0.927):.1%}")

    lp = pd.DataFrame([
        {"Metric": "Net Yield before Leverage", "Value": f"{M['yield_pre']:.2%}", "Note": "Healthy asset level return"},
        {"Metric": "Net Yield after Debt Cost", "Value": f"{M['yield_post']:.2%}", "Note": "Currently negative after leverage"},
        {"Metric": "Recognition Timing Gap", "Value": f"${M['timing_gap']/1e6:.1f}m", "Note": "Accounting versus cash"},
        {"Metric": "Net Orion Impact", "Value": f"+${M['orion_net']/1e6:.2f}m", "Note": "CDS recovery exceeded write-down"},
        {"Metric": "Major Risk Flag", "Value": "Negative levered yield + timing gap", "Note": "Fully disclosed"},
    ])
    st.dataframe(lp, use_container_width=True, hide_index=True)

    flows = pd.DataFrame({
        "Flow": ["Capital Call", "Distribution", "Call Refund", "Net"],
        "USD": [25000000, -15000000, -3200000, 6800000],
    })
    fig = px.bar(flows, x="Flow", y="USD", title="Q4 LP Capital Flows",
                 color="USD", color_continuous_scale=["#B33A3A", "#0D7377"])
    style_fig(fig, height=360)
    st.plotly_chart(fig, use_container_width=True)

    st.success(
        f"Asset level returns are solid ({M['yield_pre']:.2%}). "
        "Current levered return is temporarily negative because cost of debt exceeds loan income this quarter. "
        f"Orion was well hedged (net +${M['orion_net']/1e6:.2f}m). "
        "Two residual risks are flagged with clear owners."
    )

elif page == "Operations":
    st.markdown(
        '<div class="decision-banner">Decision: What do we fix this week — and who owns it?</div>',
        unsafe_allow_html=True,
    )
    st.title("Operations View")
    st.caption("Action queue · decision log · owners · source hierarchy")

    recs = pd.DataFrame([
        {"Priority": 1, "Action": "Restrike or unwind the IRS", "Owner": "Trading / PM",
         "Impact": f"Stops ${M['irs_basis']:,.0f} annual basis drag", "Urgency": "High"},
        {"Priority": 2, "Action": "Update JPY forward notional", "Owner": "Middle Office / FX",
         "Impact": "Removes about $2.15m residual FX exposure", "Urgency": "High"},
        {"Priority": 3, "Action": "Formalise Longwood lease policy", "Owner": "Accounting + PM",
         "Impact": f"Removes ${M['longwood']/1e6:.1f}m per year ambiguity", "Urgency": "Medium"},
        {"Priority": 4, "Action": "Tighten Greystone concentration monitoring", "Owner": "Risk",
         "Impact": "Reduces single name dependency", "Urgency": "Medium"},
        {"Priority": 5, "Action": "Monthly cash to accrual bridge", "Owner": "Accounting / Ops",
         "Impact": f"Improves visibility of ${M['timing_gap']/1e6:.1f}m timing gap", "Urgency": "Medium"},
    ])
    st.dataframe(recs, use_container_width=True, hide_index=True)

    urg = recs["Urgency"].value_counts().reset_index()
    urg.columns = ["Urgency", "Count"]
    fig = px.pie(urg, values="Count", names="Urgency", title="Actions by Urgency",
                 color_discrete_sequence=[RED, AMBER, TEAL])
    style_fig(fig, height=320)
    st.plotly_chart(fig, use_container_width=True)

    st.info(
        "The top two actions (IRS restrike and JPY forward update) address the only remaining open structural risks. "
        "Completing them would remove most of the ongoing economic drag and residual market exposure."
    )

elif page == "Scenario Lab":
    st.markdown(
        '<div class="decision-banner">Decision: How sensitive is levered yield to rates, PIK, and utilisation?</div>',
        unsafe_allow_html=True,
    )
    st.title("Scenario Lab")
    st.caption("Sidebar levers recompute yields. Base case linked from Excel.")

    g1, g2, g3, g4 = st.columns(4)
    g1.metric("Pre-Leverage Yield", f"{M['yield_pre']:.2%}")
    g2.metric("Break-even Yield", f"{M.get('be_yield', 0.1191):.2%}")
    g3.metric("Gap vs Break-even", f"{M.get('yield_gap_be', M['yield_post']):+.2%}")
    g4.metric("Post-Leverage Yield", f"{M['yield_post']:.2%}")
    st.markdown("---")

    base = {"Loan Income": M["loan_income"], "Cost of Debt": M["cost_debt"],
            "Net Income": M["net_income"], "Yield Pre": M["yield_pre"], "Yield Post": M["yield_post"]}
    scen = {"Loan Income": loan_income_s, "Cost of Debt": cost_debt_s,
            "Net Income": net_s, "Yield Pre": yield_pre_s, "Yield Post": yield_post_s}

    cmp = pd.DataFrame({
        "Metric": list(base.keys()),
        "Base Case": list(base.values()),
        "Scenario": list(scen.values()),
    })
    cmp["Delta"] = cmp["Scenario"] - cmp["Base Case"]

    st.dataframe(
        cmp.style.format({
            "Base Case": lambda x: f"{x:.2%}" if abs(x) < 1 else f"${x:,.0f}",
            "Scenario": lambda x: f"{x:.2%}" if abs(x) < 1 else f"${x:,.0f}",
            "Delta": lambda x: f"{x:+.2%}" if abs(x) < 1 else f"${x:+,.0f}",
        }),
        use_container_width=True, hide_index=True,
    )

    fig = go.Figure()
    fig.add_trace(go.Bar(name="Base", x=["Yield Pre", "Yield Post"],
                         y=[M["yield_pre"] * 100, M["yield_post"] * 100], marker_color=TEAL))
    fig.add_trace(go.Bar(name="Scenario", x=["Yield Pre", "Yield Post"],
                         y=[yield_pre_s * 100, yield_post_s * 100], marker_color=GOLD))
    style_fig(fig, title="Yield: Base vs Scenario (annualised %)", barmode="group", height=380, yaxis_title="%")
    st.plotly_chart(fig, use_container_width=True)

    fig2 = go.Figure()
    fig2.add_trace(go.Bar(name="Base", x=["Loan Income", "Cost of Debt", "Net"],
                          y=[M["loan_income"], M["cost_debt"], M["net_income"]], marker_color=TEAL))
    fig2.add_trace(go.Bar(name="Scenario", x=["Loan Income", "Cost of Debt", "Net"],
                          y=[loan_income_s, cost_debt_s, net_s], marker_color=AMBER))
    style_fig(fig2, title="Income and Cost: Base vs Scenario", barmode="group", height=380)
    st.plotly_chart(fig2, use_container_width=True)

    st.markdown("### Lever settings in force")
    st.write(f"- SOFR shock: **{sofr_bps:+d} bps**")
    st.write(f"- Extra PIK shock: **{'Yes' if pik_extra else 'No'}**")
    st.write(f"- Greystone utilisation 80%: **{'Yes' if util_up else 'No'}**")
    st.caption(
        "Method: floating share of book approximated at 65%. "
        "Facility cost moves one for one with SOFR. "
        "Extra PIK removes $1.2m cash income. "
        "Utilisation uplift adds $606k interest cost. "
        "Change levers in the sidebar and this page updates immediately."
    )

